#!/usr/bin/env python3

import sys
import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ================= CONFIG =================
INPUT_CSV  = "raw_audit_output.csv"
OUTPUT_XLS = "aws_resource_audit.xlsx"
CUTOFF     = "2026-03-10"

# 🔥 CHANGE THESE EVERY RUN
ACCOUNT_NAME = "Identity-Account"
ACCOUNT_ID   = "949501914513"

EXPECTED_COLS = [
    "Account","Region","Resource Type",
    "Resource Name","Resource ID","Creation Date"
]

# ================= STYLES =================
def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def header(cell):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F3864")
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border()

def data(cell):
    cell.border = thin_border()

# ================= HELPERS =================
def clean_dataframe(df):
    df.columns = [str(c).strip() for c in df.columns]
    df = df.fillna("")

    # Ensure all expected columns exist
    for col in EXPECTED_COLS:
        if col not in df.columns:
            df[col] = ""

    df = df[EXPECTED_COLS]
    return df

# ================= LOAD NEW DATA =================
if not os.path.exists(INPUT_CSV):
    print("❌ CSV not found")
    sys.exit(1)

new_df = pd.read_csv(INPUT_CSV, dtype=str).fillna("")
new_df.columns = EXPECTED_COLS
new_df["Account"] = ACCOUNT_NAME

# ================= LOAD EXISTING EXCEL =================
all_data = []

if os.path.exists(OUTPUT_XLS):
    wb = load_workbook(OUTPUT_XLS)

    for sheet in wb.sheetnames:
        if sheet == "Summary":
            continue

        ws = wb[sheet]
        rows = list(ws.values)

        # Skip empty or invalid sheets
        if len(rows) < 3:
            continue

        try:
            cols = list(rows[1])
            data_rows = [r for r in rows[2:] if any(r)]

            df_sheet = pd.DataFrame(data_rows, columns=cols)
            df_sheet = clean_dataframe(df_sheet)

            if not df_sheet.empty:
                all_data.append(df_sheet)

        except Exception:
            # Skip corrupted sheets silently
            continue

else:
    wb = Workbook()
    wb.remove(wb.active)

# Add new data
all_data.append(clean_dataframe(new_df))

# ================= MERGE ALL DATA =================
df = pd.concat(all_data, ignore_index=True)

df["Account"] = df["Account"].astype(str).str.strip()
df = df[df["Account"] != ""]

# ================= REMOVE OLD SHEETS =================
for sheet in wb.sheetnames:
    del wb[sheet]

# ================= WRITE ACCOUNT SHEETS =================
accounts = sorted(df["Account"].unique())

for acc in accounts:
    acc_df = df[df["Account"] == acc].reset_index(drop=True)

    ws = wb.create_sheet(acc[:31])  # Excel sheet limit

    # Title
    ws["A1"] = f"Account: {acc} ({ACCOUNT_ID}) | Before {CUTOFF} | Total: {len(acc_df)}"
    ws.merge_cells("A1:F1")

    # Header
    for i, col in enumerate(EXPECTED_COLS, 1):
        header(ws.cell(row=2, column=i, value=col))

    # Data
    for r, row in acc_df.iterrows():
        for c, col in enumerate(EXPECTED_COLS, 1):
            ws.cell(row=r+3, column=c, value=row[col])

# ================= ALL RESOURCES =================
ws_all = wb.create_sheet("All Resources")

for i, col in enumerate(EXPECTED_COLS, 1):
    header(ws_all.cell(row=1, column=i, value=col))

for r, row in df.iterrows():
    for c, col in enumerate(EXPECTED_COLS, 1):
        ws_all.cell(row=r+2, column=c, value=row[col])

# ================= SUMMARY =================
ws_sum = wb.create_sheet("Summary", 0)

ws_sum["A1"] = f"AWS Resource Audit (Before {CUTOFF})"
ws_sum.merge_cells("A1:C1")

summary = df.groupby("Account").size().reset_index(name="Count")

# Header
header(ws_sum["A3"])
header(ws_sum["B3"])
ws_sum["A3"] = "Account"
ws_sum["B3"] = "Count"

# Data
for i, row in summary.iterrows():
    ws_sum.cell(row=i+4, column=1, value=row["Account"])
    ws_sum.cell(row=i+4, column=2, value=row["Count"])

# ================= SAVE =================
try:
    wb.save(OUTPUT_XLS)
    print("✅ Excel updated:", OUTPUT_XLS)
    print("Accounts:", list(accounts))
    print("Total Resources:", len(df))
except PermissionError:
    print("❌ Close the Excel file and retry")
